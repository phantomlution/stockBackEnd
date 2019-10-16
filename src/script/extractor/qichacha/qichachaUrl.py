'''
    购买一天企查查会员，使用会员批量查询接口，获取所有的url
    来获取映射关系
        company-name -> url
'''

import os
import openpyxl
from src.service.StockService import StockService

client = StockService.getMongoInstance()
company_business_document = client.stock.company_business
base_document = client.stock.base
file_dir = os.path.dirname(__file__)


def get_excel_result(file_name):
    result = []

    file_path = file_dir + '/' + file_name
    wb = openpyxl.load_workbook(file_path)
    sheet_name_list = wb.sheetnames
    if len(sheet_name_list) != 1:
        raise Exception('数据异常：多页数据')
    sheet = wb[sheet_name_list[0]]
    row_count = sheet.max_row

    for row in range(3, row_count + 1):
        target_cell = sheet.cell(row=row, column=1)
        target = target_cell.hyperlink.target
        company_name = str.strip(target_cell.value.replace("（", '(').replace('）', ')'))
        company_url = str.strip(target.split('?')[0])

        result.append({
            "name": company_name,
            "url": company_url
        })
    return result


def extract():
    file_name = '批量查询_企业_企查查(24625244).xlsx'
    result = get_excel_result(file_name)

    for item in result:
        company_business_document.update({ "name": item['name'] }, item, True)


# 更新企业信息地址
def update_base_company_business_url():
    base_list = list(base_document.find())
    count = 0
    for base in base_list:
        company_name = base['company']['org_name_cn']
        url_item = company_business_document.find_one({ "name": company_name })
        count += 1
        if url_item is None:
            print(count)
            raise Exception('url is not: {}'.format(company_name))

        base_document.update({ "symbol": base['symbol']}, { '$set': { "qichacha_url": url_item['url']} })


if __name__ == '__main__':
    # extract()
    update_base_company_business_url()